import openpyxl
import glob
import os
import re
import shutil

root = r'C:\Users\harsh\.gemini\antigravity\scratch\vaper'
public_images = os.path.join(root, 'public', 'images')
products_js_path = os.path.join(root, 'src', 'data', 'products.js')

# ---- Step 1: Build image lookup from All images folder ----
all_imgs = []
for ext in ['*.jpg', '*.jpeg', '*.png', '*.webp']:
    all_imgs.extend(glob.glob(os.path.join(root, '**', ext), recursive=True))
image_lookup = {os.path.basename(p).lower(): p for p in all_imgs}
print(f'Total images available: {len(image_lookup)}')

# ---- Step 2: Parse all 6 new Excel files ----
def parse_price(val):
    if val is None:
        return 9.99
    s = str(val).replace('$', '').strip()
    nums = re.findall(r'[\d.]+', s)
    return round(float(nums[0]), 2) if nums else 9.99

def slugify_category(cat):
    if not cat:
        return 'other'
    cat = cat.lower()
    if 'torch' in cat or 'lighter' in cat:
        return 'lighters-torches'
    if 'grinder' in cat:
        return 'grinders'
    if 'glass' in cat or 'pipe' in cat or 'bubbler' in cat or 'bong' in cat or 'percolator' in cat:
        return 'glass-pipes'
    if 'smoking' in cat or 'wrap' in cat or 'cone' in cat or 'paper' in cat or 'filter' in cat or 'hemp' in cat:
        return 'papers-wraps'
    if 'hookah' in cat or 'charcoal' in cat or 'shisha' in cat:
        return 'hookah'
    if 'supplement' in cat or 'detox' in cat:
        return 'supplements'
    if 'candle' in cat:
        return 'accessories'
    if 'storage' in cat or 'stash' in cat or 'jar' in cat or 'pouch' in cat:
        return 'storage'
    if 'tray' in cat or 'ashtray' in cat:
        return 'accessories'
    if 'novelty' in cat or 'gift' in cat:
        return 'novelty'
    return 'other'

new_products = []

# File 14: ['Image File Name', 'Product Name', 'Estimated Retail Price']
wb = openpyxl.load_workbook(os.path.join(root, '14.xlsx'))
ws = wb.active
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[1]:
        new_products.append({
            'name': str(row[1]).strip(),
            'price': parse_price(row[2]),
            'image_file': str(row[0]).strip() if row[0] else '',
            'description': str(row[1]).strip(),
            'category_raw': 'Smoking Accessories',
        })

# File 16: ['Image Name', 'Item Description', 'Estimated Price Range']  (skip 17 — duplicates of 16)
wb = openpyxl.load_workbook(os.path.join(root, '16.xlsx'))
ws = wb.active
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[1]:
        new_products.append({
            'name': str(row[1]).strip(),
            'price': parse_price(row[2]),
            'image_file': str(row[0]).strip() if row[0] else '',
            'description': str(row[1]).strip(),
            'category_raw': 'Glass & Accessories',
        })

# File 18: ['Product Name', 'Category', 'Size / Quantity', 'Suggested Retail Price (USD)'] — no image column
wb = openpyxl.load_workbook(os.path.join(root, '18.xlsx'))
ws = wb.active
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0]:
        new_products.append({
            'name': str(row[0]).strip(),
            'price': parse_price(row[3]),
            'image_file': '',
            'description': f"{row[0]} - {row[2]}" if row[2] else str(row[0]),
            'category_raw': str(row[1]).strip() if row[1] else 'Other',
        })

# File 19: ['Item Name', 'Estimated Retail Price', 'Image Link (File Name)']
wb = openpyxl.load_workbook(os.path.join(root, '19.xlsx'))
ws = wb.active
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0]:
        new_products.append({
            'name': str(row[0]).strip(),
            'price': parse_price(row[1]),
            'image_file': str(row[2]).strip() if row[2] else '',
            'description': str(row[0]).strip(),
            'category_raw': 'Glass & Accessories',
        })

# File 20: ['Product Name', 'Estimated Pricing', 'Image Link (Filename)']
wb = openpyxl.load_workbook(os.path.join(root, '20.xlsx'))
ws = wb.active
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0]:
        new_products.append({
            'name': str(row[0]).strip(),
            'price': parse_price(row[1]),
            'image_file': str(row[2]).strip() if row[2] else '',
            'description': str(row[0]).strip(),
            'category_raw': 'Glass & Accessories',
        })

print(f'New products parsed: {len(new_products)}')

# ---- Step 3: Get current max id from products.js ----
with open(products_js_path, encoding='utf-8') as f:
    current_js = f.read()
existing_ids = re.findall(r'"id": (\d+)', current_js)
max_id = max(int(i) for i in existing_ids) if existing_ids else 108
print(f'Current max product id: {max_id}')

# ---- Step 4: Match + copy images, build JS entries ----
new_js_items = []
matched = 0
no_image = 0

for i, p in enumerate(new_products, start=max_id + 1):
    img_file = p['image_file']
    img_path = '/images/product-1.jpg'  # fallback

    if img_file:
        img_lower = img_file.lower()
        if img_lower in image_lookup:
            src = image_lookup[img_lower]
            dst = os.path.join(public_images, img_file)
            if not os.path.exists(dst):
                shutil.copy2(src, dst)
            img_path = f'/images/{img_file}'
            matched += 1
        else:
            print(f'  NOT FOUND: {img_file} for "{p["name"]}"')
            no_image += 1
    else:
        no_image += 1

    # Apply 4x pricing
    final_price = round(p['price'] * 4, 2)
    category = slugify_category(p['category_raw'])
    name_js = p['name'].replace('\\', '\\\\').replace('"', '\\"')
    desc_js = p['description'].replace('\\', '\\\\').replace('"', '\\"')
    cat_raw_js = p['category_raw'].replace('"', '\\"')

    new_js_items.append(f'''  {{
    "id": {i},
    "name": "{name_js}",
    "category": "{category}",
    "categoryLabel": "{cat_raw_js}",
    "price": {final_price},
    "salePrice": null,
    "description": "{desc_js}",
    "image": "{img_path}",
    "badge": "new",
    "inStock": true,
    "featured": false
  }}''')

print(f'Images matched: {matched}, No image: {no_image}')
print(f'New entries to add: {len(new_js_items)}')

# ---- Step 5: Append to products.js ----
# Find the closing ]; of products array and insert before it
new_block = ',\n' + ',\n'.join(new_js_items)
updated_js = re.sub(r'(\];(\s*\nexport const categories))', new_block + r'\n\1', current_js, count=1)

with open(products_js_path, 'w', encoding='utf-8') as f:
    f.write(updated_js)

# Verify
total_ids = re.findall(r'"id": (\d+)', updated_js)
print(f'Total products in products.js: {len(total_ids)}')
print('DONE!')
