import openpyxl
import glob
import os

files = glob.glob(r'C:\Users\harsh\.gemini\antigravity\scratch\vaper\*.xlsx')
for f in files:
    print('='*60)
    print('FILE:', f)
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    headers = [str(cell.value) for cell in ws[1]]
    print('HEADERS:', headers)
    print('ROWS:', ws.max_row)

    # Check for hyperlinks in cells
    print('\n--- HYPERLINKS IN FIRST 5 ROWS ---')
    for row in ws.iter_rows(min_row=2, max_row=6):
        for cell in row:
            if cell.hyperlink:
                print(f'  Cell {cell.coordinate}: value={cell.value!r}, hyperlink={cell.hyperlink.target!r}')

    # Print all row values for small file
    if os.path.getsize(f) < 50000:
        print('\n--- ALL ROWS ---')
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 1):
            if any(v is not None for v in row):
                row_dict = dict(zip(headers, row))
                print(f"Row {i}: Name={row_dict.get('Item Name')!r} | Price={row_dict.get('Estimated Price')!r} | Cat={row_dict.get('Product Category')!r} | ImgFile={row_dict.get('Image_File')!r} | ImgLink={row_dict.get('Image_Link')!r}")

    # Check for embedded images
    print('\n--- EMBEDDED IMAGES ---')
    if hasattr(ws, '_images'):
        print(f'Found {len(ws._images)} embedded images')
        for i, img in enumerate(ws._images[:3]):
            print(f'  Image {i}: anchor={img.anchor}, format={getattr(img, "format", "?")}')
    else:
        print('No _images attribute')
