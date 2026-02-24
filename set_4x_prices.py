import re
import openpyxl
import glob
import os

# Read original Excel prices
files = glob.glob(r'C:\Users\harsh\.gemini\antigravity\scratch\vaper\*.xlsx')
target_xlsx = [f for f in files if os.path.getsize(f) < 50000][0]
wb = openpyxl.load_workbook(target_xlsx)
ws = wb.active
headers = [str(cell.value) for cell in ws[1]]
excel_products = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if any(v is not None for v in row):
        excel_products.append(dict(zip(headers, row)))

def parse_price(price_str):
    if not price_str:
        return 9.99
    price_str = str(price_str).replace('$', '').strip()
    nums = re.findall(r'[\d.]+', price_str)
    return round(float(nums[0]), 2) if nums else 9.99

# Load current products.js
with open(r'C:\Users\harsh\.gemini\antigravity\scratch\vaper\src\data\products.js', encoding='utf-8') as f:
    content = f.read()

# Replace all "price": <value> using original Excel prices * 4
# Build list of (id, new_price)
price_map = {}
for i, p in enumerate(excel_products, start=1):
    orig = parse_price(p.get('Estimated Price', ''))
    price_map[i] = round(orig * 4, 2)

# Replace prices one by one by product id block
for pid, new_price in price_map.items():
    # Match the id block then find its price field
    pattern = r'("id": ' + str(pid) + r',\s*"name":.*?"price": )[\d.]+'
    replacement = r'\g<1>' + str(new_price)
    content = re.sub(pattern, replacement, content, count=1, flags=re.DOTALL)

with open(r'C:\Users\harsh\.gemini\antigravity\scratch\vaper\src\data\products.js', 'w', encoding='utf-8') as f:
    f.write(content)

prices = re.findall(r'"price": ([\d.]+)', content)
print('Updated products:', len(prices))
print('First 5 new prices (4x original):', prices[:5])
# Verify: product 1 original was $8.99 -> 4x = $35.96
print('Product 1 expected: $35.96, got: $' + prices[0])
