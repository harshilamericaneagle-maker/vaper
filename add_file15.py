import openpyxl
import re
import os
import glob
import shutil

root = r'C:\Users\harsh\.gemini\antigravity\scratch\vaper'
public_images = os.path.join(root, 'public', 'images')
products_js_path = os.path.join(root, 'src', 'data', 'products.js')

# Build image lookup
all_imgs = []
for ext in ['*.jpg', '*.jpeg', '*.png', '*.webp']:
    all_imgs.extend(glob.glob(os.path.join(root, '**', ext), recursive=True))
image_lookup = {os.path.basename(p).lower(): p for p in all_imgs}

# Already-used images in public/images
used_images = set(os.path.basename(f).lower() for f in glob.glob(os.path.join(public_images, '*.jpg')))
print(f'Images available: {len(image_lookup)}, Already used: {len(used_images)}')

# Get unused images (candidates for products without image column)
unused_images = [k for k in image_lookup.keys() if k not in used_images and k.startswith('2026')]
print(f'Unused candidate images: {len(unused_images)}')

def parse_price(val):
    if val is None:
        return 9.99
    s = str(val).replace('$', '').replace('\û', '-').strip()
    nums = re.findall(r'[\d.]+', s)
    return round(float(nums[0]), 2) if nums else 9.99

def slugify_category(name, feat=''):
    t = (name + ' ' + str(feat)).lower()
    if 'candle' in t or 'odor' in t:
        return 'accessories'
    if 'pipe' in t or 'bong' in t or 'bubbler' in t or 'bowl' in t or 'mouthpiece' in t or 'drip tip' in t:
        return 'glass-pipes'
    if 'grinder' in t:
        return 'grinders'
    if 'silicone' in t:
        return 'glass-pipes'
    if 'connector' in t or 'wire' in t:
        return 'accessories'
    return 'other'

# Parse 15.xlsx
wb = openpyxl.load_workbook(os.path.join(root, '15.xlsx'))
ws = wb.active
headers = [str(c.value) for c in ws[1]]
new_products = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0]:
        new_products.append({
            'name': str(row[0]).strip(),
            'price': parse_price(row[1]),
            'features': str(row[2]).strip() if row[2] else '',
        })
print(f'Products from 15.xlsx: {len(new_products)}')

# Get current max id
with open(products_js_path, encoding='utf-8') as f:
    current_js = f.read()
existing_ids = re.findall(r'"id": (\d+)', current_js)
max_id = max(int(i) for i in existing_ids) if existing_ids else 156
print(f'Current max id: {max_id}')

# Build JS entries
new_js_items = []
for i, p in enumerate(new_products, start=max_id + 1):
    # Try to find an unused image to assign
    img_path = '/images/product-' + str(((i - 1) % 5) + 1) + '.jpg'  # fallback
    if unused_images:
        img_file = unused_images.pop(0)
        src = image_lookup[img_file]
        dst = os.path.join(public_images, img_file)
        if not os.path.exists(dst):
            shutil.copy2(src, dst)
        img_path = f'/images/{img_file}'

    final_price = round(p['price'] * 4, 2)
    category = slugify_category(p['name'], p['features'])
    name_js = p['name'].replace('\\', '\\\\').replace('"', '\\"')
    desc_js = p['features'].replace('\\', '\\\\').replace('"', '\\"') if p['features'] else name_js

    new_js_items.append(f'''  {{
    "id": {i},
    "name": "{name_js}",
    "category": "{category}",
    "categoryLabel": "New Arrivals",
    "price": {final_price},
    "salePrice": null,
    "description": "{desc_js}",
    "image": "{img_path}",
    "badge": "new",
    "inStock": true,
    "featured": false
  }}''')

print(f'New JS entries built: {len(new_js_items)}')

# Append to products.js
new_block = ',\n' + ',\n'.join(new_js_items)
updated_js = re.sub(r'(\];(\s*\nexport const categories))', new_block + r'\n\1', current_js, count=1)

with open(products_js_path, 'w', encoding='utf-8') as f:
    f.write(updated_js)

total_ids = re.findall(r'"id": (\d+)', updated_js)
print(f'Total products now: {len(total_ids)}')
print('DONE!')
