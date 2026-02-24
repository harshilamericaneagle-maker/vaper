import openpyxl, glob, json, os

files = glob.glob(r'C:\Users\harsh\.gemini\antigravity\scratch\vaper\*.xlsx')
for f in files:
    print('='*60)
    print('FILE:', f)
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    headers = [str(cell.value) for cell in ws[1]]
    print('HEADERS:', headers)
    print('TOTAL ROWS:', ws.max_row)
    print()
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=1):
        if i <= 5:
            print(f'Row {i}:', list(row))

# Also check the new smaller excel - read ALL rows for that one
for f in files:
    size = os.path.getsize(f)
    if size < 50000:  # smaller file = the 108 products one
        print('\n\n### SMALL FILE - ALL ROWS ###')
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        headers = [str(cell.value) for cell in ws[1]]
        print('HEADERS:', headers)
        all_data = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if any(v is not None for v in row):
                all_data.append(dict(zip(headers, row)))
        print(f'Total products: {len(all_data)}')
        print()
        for item in all_data[:10]:
            print(item)
