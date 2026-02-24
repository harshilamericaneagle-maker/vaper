import openpyxl
import glob
import os
import re

# ---- Step 1: Read Excel ----
xlsx_files = glob.glob(r'C:\Users\harsh\.gemini\antigravity\scratch\vaper\*.xlsx')
target_xlsx = [f for f in xlsx_files if os.path.getsize(f) < 50000][0]
wb = openpyxl.load_workbook(target_xlsx)
ws = wb.active
headers = [str(cell.value) for cell in ws[1]]
excel_products = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if any(v is not None for v in row):
        excel_products.append(dict(zip(headers, row)))

# ---- Step 2: Category slug mapping ----
def slugify_category(cat):
    if not cat:
        return 'other'
    cat = cat.lower()
    if 'torch' in cat or 'lighter' in cat:
        return 'lighters-torches'
    if 'grinder' in cat:
        return 'grinders'
    if 'glass' in cat or 'pipe' in cat or 'bubbler' in cat:
        return 'glass-pipes'
    if 'smoking' in cat or 'wrap' in cat or 'cone' in cat or 'paper' in cat or 'filter' in cat or 'hemp' in cat:
        return 'papers-wraps'
    if 'hookah' in cat or 'charcoal' in cat or 'shisha' in cat:
        return 'hookah'
    if 'supplement' in cat or 'detox' in cat:
        return 'supplements'
    if 'snack' in cat or 'consumable' in cat:
        return 'other'
    if 'novelty' in cat or 'gift' in cat:
        return 'novelty'
    if 'storage' in cat or 'smell' in cat or 'pouch' in cat:
        return 'storage'
    if 'scale' in cat:
        return 'accessories'
    if 'defense' in cat or 'pepper' in cat:
        return 'accessories'
    if 'dispenser' in cat or 'accessori' in cat:
        return 'accessories'
    if 'home' in cat:
        return 'accessories'
    return 'other'

# ---- Step 3: Parse price range (take average or low price as display price) ----
def parse_price(price_str):
    if not price_str:
        return 9.99
    price_str = str(price_str).replace('$', '').strip()
    prices = re.findall(r'[\d.]+', price_str)
    if len(prices) >= 2:
        return round(float(prices[0]), 2)   # use the lower price
    elif len(prices) == 1:
        return round(float(prices[0]), 2)
    return 9.99

# ---- Step 4: Build products array ----
categories_seen = set()
products_js_items = []

for idx, p in enumerate(excel_products, start=1):
    name = str(p.get('Item Name', '') or '').strip()
    price_raw = p.get('Estimated Price', '') or ''
    description = str(p.get('Brief Description', '') or '').strip()
    category_raw = str(p.get('Product Category', '') or '').strip()
    img_file = str(p.get('Image_File', '') or '').strip()
    
    price = parse_price(price_raw)
    category = slugify_category(category_raw)
    categories_seen.add((category, category_raw))
    
    # Escape special chars for JS
    name_js = name.replace('\\', '\\\\').replace('"', '\\"').replace("'", "\\'")
    desc_js = description.replace('\\', '\\\\').replace('"', '\\"').replace("'", "\\'")
    cat_raw_js = category_raw.replace('\\', '\\\\').replace('"', '\\"')

    img_path = f'/images/{img_file}' if img_file else '/images/product-1.jpg'
    
    featured = 'true' if idx <= 12 else 'false'
    
    item = f'''  {{
    "id": {idx},
    "name": "{name_js}",
    "category": "{category}",
    "categoryLabel": "{cat_raw_js}",
    "price": {price},
    "salePrice": null,
    "description": "{desc_js}",
    "image": "{img_path}",
    "badge": {"\"new\"" if idx <= 20 else "null"},
    "inStock": true,
    "featured": {featured}
  }}'''
    products_js_items.append(item)

# ---- Step 5: Build categories array from what's in the data ----
# Get all unique slugs + their labels
slug_to_label = {}
for slug, raw in categories_seen:
    # Pick a nice label
    if slug not in slug_to_label:
        slug_to_label[slug] = raw

# Nice human-readable labels for each slug
nice_labels = {
    'lighters-torches': 'Lighters & Torches',
    'grinders': 'Grinders',
    'glass-pipes': 'Glass & Pipes',
    'papers-wraps': 'Papers & Wraps',
    'hookah': 'Hookah',
    'supplements': 'Supplements',
    'novelty': 'Novelty & Gifts',
    'storage': 'Storage',
    'accessories': 'Accessories',
    'other': 'Other',
}

categories_js_items = [
    '  { "id": "all", "name": "All Products" }',
    '  { "id": "sale", "name": "On Sale" }',
]
for slug in sorted(slug_to_label.keys()):
    label = nice_labels.get(slug, slug.replace('-', ' ').title())
    categories_js_items.append(f'  {{ "id": "{slug}", "name": "{label}" }}')

# ---- Step 6: Write new products.js ----
output = '''// Product data - Auto-generated from Final 108 products Excel
// All 108 products with real images, names, descriptions, and prices

export const products = [
''' + ',\n'.join(products_js_items) + '''
];

export const categories = [
''' + ',\n'.join(categories_js_items) + '''
];
'''

out_path = r'C:\Users\harsh\.gemini\antigravity\scratch\vaper\src\data\products.js'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(output)

print(f"Written {len(excel_products)} products to products.js")
print(f"Categories ({len(categories_js_items)}):")
for c in categories_js_items:
    print(f"  {c}")
print(f"\nFile size: {os.path.getsize(out_path):,} bytes")
print("DONE!")
