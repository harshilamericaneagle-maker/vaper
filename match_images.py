import openpyxl
import glob
import os
import shutil
import json

# --- Step 1: Find the Excel file with 108 products ---
xlsx_files = glob.glob(r'C:\Users\harsh\.gemini\antigravity\scratch\vaper\*.xlsx')
target_xlsx = [f for f in xlsx_files if os.path.getsize(f) < 50000][0]
print(f"Using Excel: {target_xlsx}")

# --- Step 2: Read Excel data ---
wb = openpyxl.load_workbook(target_xlsx)
ws = wb.active
headers = [str(cell.value) for cell in ws[1]]
excel_products = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if any(v is not None for v in row):
        d = dict(zip(headers, row))
        excel_products.append(d)
print(f"Excel products: {len(excel_products)}")

# --- Step 3: Find all jpg/png images recursively ---
root = r'C:\Users\harsh\.gemini\antigravity\scratch\vaper'
all_images = []
for ext in ['*.jpg', '*.jpeg', '*.png', '*.webp']:
    all_images.extend(glob.glob(os.path.join(root, '**', ext), recursive=True))

# Build a lookup dict by filename (lowercase)
image_lookup = {}
for img_path in all_images:
    fname = os.path.basename(img_path).lower()
    image_lookup[fname] = img_path

print(f"\nTotal images found: {len(all_images)}")
print("\nSample images found:")
for k, v in list(image_lookup.items())[:10]:
    print(f"  {k} -> {v}")

# --- Step 4: Match Excel image filenames to actual files and copy to public/images ---
public_images_dir = r'C:\Users\harsh\.gemini\antigravity\scratch\vaper\public\images'
os.makedirs(public_images_dir, exist_ok=True)

matched = 0
unmatched = []
for product in excel_products:
    img_file = str(product.get('Image_File', '') or '').strip()
    if not img_file:
        unmatched.append(product.get('Item Name', ''))
        continue
    img_file_lower = img_file.lower()
    if img_file_lower in image_lookup:
        src = image_lookup[img_file_lower]
        dst = os.path.join(public_images_dir, img_file)
        if not os.path.exists(dst):
            shutil.copy2(src, dst)
        matched += 1
    else:
        unmatched.append(f"{product.get('Item Name')} (img: {img_file})")

print(f"\nMatched images: {matched}")
print(f"Unmatched: {len(unmatched)}")
if unmatched:
    for u in unmatched[:10]:
        print(f"  UNMATCHED: {u}")
