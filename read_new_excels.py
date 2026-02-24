import openpyxl
import glob
import os

new_files = ['14.xlsx', '16.xlsx', '17.xlsx', '18.xlsx', '19.xlsx', '20.xlsx']
root = r'C:\Users\harsh\.gemini\antigravity\scratch\vaper'

for fname in new_files:
    fpath = os.path.join(root, fname)
    wb = openpyxl.load_workbook(fpath)
    ws = wb.active
    headers = [str(cell.value) for cell in ws[1]]
    print(f'\n=== {fname} ===')
    print(f'Headers: {headers}')
    print(f'Rows: {ws.max_row}')
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if any(v is not None for v in row):
            print(dict(zip(headers, row)))
